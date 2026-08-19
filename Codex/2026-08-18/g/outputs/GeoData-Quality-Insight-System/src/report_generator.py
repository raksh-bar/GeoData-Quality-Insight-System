"""Excel and chart output for the project."""

from pathlib import Path

import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_DICTIONARY = [
    ("place_id", "Unique identifier for the location", "Required"),
    ("place_name", "Name of the place or point of interest", "Required"),
    ("city", "City in which the place is located", "Required"),
    ("latitude", "North/south coordinate; valid range is -90 to 90", "Required"),
    ("longitude", "East/west coordinate; valid range is -180 to 180", "Required"),
    ("normal_eta_min", "Expected travel time in minutes; must be greater than zero", "Required"),
    ("actual_eta_min", "Observed travel time in minutes", "Required"),
    ("defect_type", "All detected data-quality issues for the record", "Generated"),
    ("severity", "Business priority of the most serious issue", "Generated"),
]


def _format_workbook(writer: pd.ExcelWriter) -> None:
    """Make every worksheet readable without technical knowledge."""
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            widest = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(widest + 3, 55)
        worksheet.row_dimensions[1].height = 30


def create_excel_report(
    all_records: pd.DataFrame,
    summary: dict[str, object],
    city_summary: pd.DataFrame,
    type_summary: pd.DataFrame,
    findings: list[dict[str, str]],
    duplicates: pd.DataFrame,
    eta_issues: pd.DataFrame,
    output_file: Path,
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    summary_df = pd.DataFrame(summary.items(), columns=["Metric", "Value"])
    guide_df = pd.DataFrame([
        ("Report purpose", "A plain-English review of location and travel-time data quality."),
        ("How to read severity", "Critical = fix immediately; High = fix before operational use; Medium = review; Valid = no issue found."),
        ("Data-quality score", "Percentage of records that have no detected issue. A higher percentage is better."),
        ("Important note", "This report flags possible issues. Confirm business context before deleting or changing a record."),
    ], columns=["Section", "Description"])
    dictionary_df = pd.DataFrame(DATA_DICTIONARY, columns=["Column", "Meaning", "Status"])
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        guide_df.to_excel(writer, sheet_name="Read Me First", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(findings).to_excel(writer, sheet_name="Key Findings", index=False)
        dictionary_df.to_excel(writer, sheet_name="Data Dictionary", index=False)
        all_records.to_excel(writer, sheet_name="All Records", index=False)
        city_summary.to_excel(writer, sheet_name="Defects by City", index=False)
        type_summary.to_excel(writer, sheet_name="Defects by Type", index=False)
        duplicates.to_excel(writer, sheet_name="Duplicates", index=False)
        eta_issues.to_excel(writer, sheet_name="ETA Anomalies", index=False)
        _format_workbook(writer)


def _font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    """Use Windows' Arial where available; safely fall back if it is not."""
    names = ["C:/Windows/Fonts/arialbd.ttf"] if bold else ["C:/Windows/Fonts/arial.ttf"]
    names += ["DejaVuSans-Bold.ttf"] if bold else ["DejaVuSans.ttf"]
    for name in names:
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            pass
    return ImageFont.load_default()


def _centered(draw: ImageDraw.ImageDraw, text: str, y: int, font: ImageFont.ImageFont, fill: str, width: int) -> None:
    box = draw.textbbox((0, 0), text, font=font)
    draw.text(((width - (box[2] - box[0])) / 2, y), text, font=font, fill=fill)


def create_dashboard(city_summary: pd.DataFrame, type_summary: pd.DataFrame, summary: dict[str, object], output_file: Path) -> None:
    """Create a company-presentation dashboard without a heavyweight chart dependency."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    width, height = 1800, 1000
    image = Image.new("RGB", (width, height), "#f7f9fc")
    draw = ImageDraw.Draw(image)
    navy, slate, orange, blue = "#17365D", "#4a5568", "#e76f51", "#457b9d"
    _centered(draw, "GeoData Quality Insight Dashboard", 32, _font(48, True), navy, width)
    _centered(draw, "A clear overview of map-data reliability and the actions needed to improve it", 96, _font(24), slate, width)

    left_x, right_x, chart_y, chart_height, chart_width = 80, 940, 260, 480, 700
    draw.text((left_x, 190), "Where are the issues?", font=_font(30, True), fill="#111111")
    draw.text((right_x, 190), "What needs attention?", font=_font(30, True), fill="#111111")
    draw.rectangle((left_x, chart_y, left_x + chart_width, chart_y + chart_height), fill="white")
    draw.rectangle((right_x, chart_y, right_x + chart_width, chart_y + chart_height), fill="white")

    city_items = list(zip(city_summary["city"].fillna("Unknown"), city_summary["defect_count"]))
    city_max = max((count for _, count in city_items), default=1)
    bar_space = chart_width / max(len(city_items), 1)
    for index, (city, count) in enumerate(city_items):
        bar_width = min(110, int(bar_space * 0.55))
        x = int(left_x + index * bar_space + (bar_space - bar_width) / 2)
        bar_height = int((count / city_max) * (chart_height - 95))
        y = chart_y + chart_height - 45 - bar_height
        draw.rectangle((x, y, x + bar_width, chart_y + chart_height - 45), fill=orange)
        value_box = draw.textbbox((0, 0), str(count), font=_font(22, True))
        draw.text((x + (bar_width - (value_box[2] - value_box[0])) / 2, y - 32), str(count), font=_font(22, True), fill="#111111")
        city_box = draw.textbbox((0, 0), str(city), font=_font(19))
        draw.text((x + (bar_width - (city_box[2] - city_box[0])) / 2, chart_y + chart_height - 32), str(city), font=_font(19), fill="#111111")
    draw.text((left_x + 250, chart_y + chart_height + 15), "City", font=_font(21), fill=slate)

    type_items = list(zip(type_summary["defect_type"], type_summary["count"]))
    type_max = max((count for _, count in type_items), default=1)
    row_height = (chart_height - 40) / max(len(type_items), 1)
    for index, (issue, count) in enumerate(type_items):
        y = int(chart_y + 22 + index * row_height)
        label = str(issue)
        draw.text((right_x + 16, y), label, font=_font(21), fill="#111111")
        bar_x = right_x + 275
        bar_width = int((count / type_max) * 350)
        draw.rectangle((bar_x, y + 2, bar_x + bar_width, y + 28), fill=blue)
        draw.text((bar_x + bar_width + 12, y - 1), str(count), font=_font(21, True), fill="#111111")

    score_text = (f"DATA QUALITY SCORE: {summary['Data quality percentage']}%     |     "
                  f"VALID: {summary['Valid records']:,}     |     "
                  f"REQUIRING REVIEW: {summary['Defective records']:,}")
    box = draw.textbbox((0, 0), score_text, font=_font(26, True))
    score_x = (width - (box[2] - box[0])) / 2
    draw.rounded_rectangle((score_x - 28, 825, score_x + (box[2] - box[0]) + 28, 890), radius=18, fill="#d9eaf7", outline="#9cc3d5", width=3)
    draw.text((score_x, 842), score_text, font=_font(26, True), fill=navy)
    _centered(draw, "Recommendation: fix critical coordinate errors first, then review duplicate locations and ETA differences.", 915, _font(20), slate, width)
    image.save(output_file)
